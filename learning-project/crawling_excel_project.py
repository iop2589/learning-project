import requests, openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, colors, Border, Side, PatternFill

# excel style 변수
wrap_alignment = Alignment(wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="center")
bold_font = Font(bold=True)
header_font = Font(bold=True, size=14)
hyperlink_font = Font(color=colors.BLUE, underline="single")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# excel file 변수
excel_file = openpyxl.Workbook()
excel_sheet = excel_file.active
excel_sheet.title = "Reporting to Website"

# 게시글 가져오기
res = requests.get("https://davelee-fun.github.io/trial/board/news.html")
soup = BeautifulSoup(res.content, "html.parser")

items = soup.select("div.list_item")

index = 1
row_index = 2;

excel_sheet.append([""] + ["No", "제목", "댓글"])

for cell in excel_sheet[1][1:]:
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border
    cell.fill = header_fill

for item in items[:7]:
    if item.select_one("span.subject_fixed") != None:
        title = item.select_one("span.subject_fixed").get_text().strip()
        coment_count = item.select_one("span.rSymph05").get_text().strip()
        link = item.select_one("a.list_subject")
        link_url = "https://davelee-fun.github.io/trial/board/" + link["href"]

        xl_info = [""] + [index, title, "[" + coment_count + "]"]
        excel_sheet.append(xl_info)
        
        for cell in excel_sheet[row_index][1:]:
            cell.border = thin_border
            cell.alignment = wrap_alignment
            if cell.column == 3: 
                cell.hyperlink = link_url
                cell.font = hyperlink_font
            else:
                cell.font = bold_font
            
        row_index += 1
        
        # 게시글의 댓글 가져오기
        rep_title = requests.get(link_url, "html.parser")
        rep_soup = BeautifulSoup(rep_title.content, "html.parser")
        replies = rep_soup.select("div.comment_content")

        for re_ind, reply in enumerate(replies):
            comment_content = str(re_ind + 1) + ". " + reply.select_one("div.comment_view").get_text().strip().replace("\n", "").replace("\t", "")
            xl_info = [""] + ["", "", comment_content]
            excel_sheet.append(xl_info)
            
            for cell in excel_sheet[row_index][1:]:
                cell.border = thin_border
                cell.alignment = wrap_alignment
            
            row_index += 1

        index += 1

# excel save logic

# excel colomun width 조정
excel_sheet.column_dimensions["B"].width = "20"
excel_sheet.column_dimensions["C"].width = "50"
excel_sheet.column_dimensions["D"].width = "50"

excel_file.save("xl_data_report.xlsx")
excel_file.close()